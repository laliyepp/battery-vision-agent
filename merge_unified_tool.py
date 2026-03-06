#!/usr/bin/env python3
"""Merge all *_unified_output.json files under a vehicle output directory into a single JSON array + Excel."""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

DOMAINS = ["通用", "电池", "电机", "电控"]

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)


def load_schema_keys(schema_path: Path) -> dict[str, list[str]]:
    """Load unified schema and return {domain: [ordered_keys]}."""
    with open(schema_path, encoding="utf-8") as f:
        raw = json.load(f)
    return {domain: list(fields.keys()) for domain, fields in raw.items()}


def write_excel(records: list[dict], output_path: Path, schema_keys: dict[str, list[str]]):
    """Write merged records to a 4-sheet Excel file (one sheet per domain)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for domain in DOMAINS:
        ws = wb.create_sheet(title=domain)
        keys = schema_keys.get(domain, [])
        headers = ["来源文件"] + keys

        # Write header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

        # Write data rows
        for row_idx, rec in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=rec.get("_source_file", "")).alignment = CELL_ALIGN
            domain_data = rec.get(domain, {})
            for col_idx, key in enumerate(keys, 2):
                val = domain_data.get(key, "")
                ws.cell(row=row_idx, column=col_idx, value=val).alignment = CELL_ALIGN

        # Auto-width (capped at 40)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row_idx in range(2, len(records) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value or ""
                max_len = max(max_len, min(len(str(val)), 40))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Merge unified output files into JSON + Excel")
    parser.add_argument("vehicle_dir", help="Vehicle output directory to scan recursively")
    parser.add_argument("--sort-mode", choices=["date", "history"], default="date",
                        help="Sort mode: 'date' (default) sorts by 签发日期 with declared last; "
                             "'history' sorts by subfolder ID from path then by 签发日期")
    args = parser.parse_args()

    vehicle_dir = Path(args.vehicle_dir)
    if not vehicle_dir.is_dir():
        print(f"[Error] Directory not found: {vehicle_dir}", file=sys.stderr)
        sys.exit(1)

    # Find all *_unified_output.json, exclude unified_schema_result.json
    all_files = sorted(vehicle_dir.rglob("*_unified_output.json"))
    if not all_files:
        print("[Error] No *_unified_output.json files found.", file=sys.stderr)
        sys.exit(1)

    # Separate declared (三电参数) from report files
    declared_file = None
    report_files = []
    for fp in all_files:
        if fp.name == "三电参数_unified_output.json":
            declared_file = fp
        else:
            report_files.append(fp)

    # Load all report files
    records = []
    for fp in report_files:
        with open(fp, encoding="utf-8") as f:
            data = json.load(f)
        data["_source_file"] = str(fp.relative_to(vehicle_dir))
        records.append(data)

    if args.sort_mode == "history":
        # Sort by subfolder ID (between first/ and second/ in _source_file), then by 签发日期
        def _extract_subfolder_sort_key(source_file: str):
            """Extract subfolder ID from path like 'history/20190911010892B10/...' and return a natural sort key."""
            parts = source_file.split("/")
            subfolder = parts[1] if len(parts) >= 2 else ""
            # Pattern: digits + B/K + number + optional S + number
            # e.g. '20190911010892B13S1' -> ('20190911010892', 'B', 13, 'S1')
            m = re.match(r"^(\d+)([BK])(\d+)(S\d+)?$", subfolder)
            if m:
                return (m.group(1), m.group(2), int(m.group(3)), m.group(4) or "")
            return (subfolder, "", 0, "")

        def sort_key_history(rec):
            source = rec.get("_source_file", "")
            date = rec.get("通用", {}).get("签发日期", "")
            return (_extract_subfolder_sort_key(source), date)

        records.sort(key=sort_key_history)
    else:
        # Default: sort by 通用.签发日期 ascending (empty dates sort first)
        def sort_key_date(rec):
            date = rec.get("通用", {}).get("签发日期", "")
            return (0 if not date else 1, date)

        records.sort(key=sort_key_date)

    # Append declared file last (only in date mode; in history mode there's no declared file)
    if declared_file:
        with open(declared_file, encoding="utf-8") as f:
            declared_data = json.load(f)
        declared_data["_source_file"] = str(declared_file.relative_to(vehicle_dir))
        records.append(declared_data)

    # Write merged JSON
    json_path = vehicle_dir / "merged_final_result.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    # Write merged Excel
    schema_path = Path(__file__).parent / "schema" / "unified_schema_v2.json"
    schema_keys = load_schema_keys(schema_path)
    xlsx_path = vehicle_dir / "merged_final_result.xlsx"
    write_excel(records, xlsx_path, schema_keys)

    n_reports = len(report_files)
    n_declared = 1 if declared_file else 0
    total = len(records)
    print(f"[Done] merged_final_result.json — {total} records ({n_reports} reports + {n_declared} declared)")
    print(f"[Done] merged_final_result.xlsx — 4 sheets ({', '.join(DOMAINS)})")


if __name__ == "__main__":
    main()
